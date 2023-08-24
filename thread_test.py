from threading import Thread, Event, Lock
from queue import Queue
from time import sleep, time
import datetime
from datetime import date
from openpyxl import Workbook,load_workbook
from kotak_online import kotak
import os
import random
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

import logging


data_lock = Lock()
event = Event()
globalVar = 0
q1 = Queue()
q2 = Queue()


class DUMMY_KOTAK():
    def __init__(self) -> None:
        
        # Set logger 
        today = date.today()
        d = today.strftime("%Y%m%d")
        log_path = "/home/administrator/project/stockbot/logs/"+d
        log_path_file = log_path+"/app.log"
        
        try:
            os.makedirs(log_path) 
            
        except Exception as e:
            print(e)
            if FileExistsError:              
                pass
            else:
                print(f"Log Path creation failed- {log_path}")
        logger = logging.getLogger(__name__)
        logging.basicConfig(level=logging.DEBUG, format='%(asctime)s %(name)-12s %(levelname)-8s %(message)s')

        LOG_FILENAME=log_path_file
        #Define handler to write to standard output
        handler = logging.FileHandler(LOG_FILENAME)

        formatter = logging.Formatter('[%(levelname)s] %(message)s')
        handler.setFormatter(formatter)
        handler.setLevel(logging.DEBUG)

        #Adding handler to logger
        logger.addHandler(handler)
        logger.propagate=False
        self.logger =logger
        
        #Define logger


        
        
        
        
        
        
        
        
        
        
        
        
        
        self.sleep_time = 0
        self.q_case = 1
        self.q_print = 0
        self.count = 0
        self.maxQ = 50
        self.kill = False
        self.kk = kotak(
            path_ini='/home/administrator/git/stock_bot/config.ini', logger=self.logger)
        self.dump_threshold = self.maxQ
        self.csv_dump_path = "/home/administrator/project/stockbot/daily_dumps/"+str(d)+".xlsx"
        # self.csv_dump_path = "/home/administrator/project/stockbot/daily_dumps/depth_csv_dump.xlsx"
        self.createXLXS()
        
        
        
        
        
        
    def callBackStock(self,):
        while (True):  
            self.count = self.count+1
            if (self.count % 2) ==0:
                print(f'count: {self.count}')
            if (self.q_case == 1):
                if (q1.qsize() < self.maxQ):
                    with data_lock:
                        q1.put( e.kk.client.quote(instrument_token = 1900,quote_type='DEPTH'))
                else:
                    with data_lock:
                        self.q_case = 2
                        self.q_print = 1
            else:
                if (q2.qsize() < self.maxQ):
                    with data_lock:
                        q2.put( e.kk.client.quote(instrument_token = 1900,quote_type='DEPTH'))

                else:
                    with data_lock:
                        self.q_case = 1
                        self.q_print = 2
                        
            if event.is_set():
                break
            else:
                sleep(0.1)
            

    def random_sleep(self,):
        r1 = random.randint(1, 100)
        print("sleep for ", str(r1/1000))
        return r1/100

    def send_fake_stock(self,):
        while (not self.kill):
            sleep(self.random_sleep())

            if (self.q_case == 1):
                if (q1.qsize() < self.maxQ):
                    with data_lock:
                        q1.put(str(random.randint(1000, 9999)))
                        # print("@q1")
                else:
                    with data_lock:
                        self.q_case = 2
                        self.q_print = 1
            else:
                if (q2.qsize() < self.maxQ):
                    with data_lock:
                        q2.put(str(random.randint(1000, 9999)))
                        # print("@q2")
                else:
                    with data_lock:
                        self.q_case = 1
                        self.q_print = 2

            if event.is_set():
                break

    def stock_call_back(self, msg):
        if (self.count % 50) ==0:
            print(f'count: {self.count}')
        self.count = self.count+1
        if (self.q_case == 1):
            if (q1.qsize() < self.maxQ):
                with data_lock:
                    q1.put(msg)
            else:
                with data_lock:
                    self.q_case = 2
                    self.q_print = 1
        else:
            if (q2.qsize() < self.maxQ):
                with data_lock:
                    q2.put(msg)
                    # print("@q2")
            else:
                with data_lock:
                    self.q_case = 1
                    self.q_print = 2

    
    def avoid_duplicates(self,msg):
        if (self.q_case == 1):
            
            pass
        else:
            
            pass
    
    
    def print_q(self,):
        data=[]
        while (True):
            if (self.q_print > 0):
                if(self.q_print == 2 and q2.qsize()==self.maxQ):
                    for i in range(q2.qsize()):
                        # print(q2.get(i))
                        data.append(q2.get(i))
                        # sleep(0.1)
                    q2.queue.clear()
                    
                    
                elif(self.q_print == 1 and q1.qsize()==self.maxQ):
                    for i in range(q1.qsize()):
                        # print(q1.get(i))
                        data.append(q1.get(i))
                        # sleep(0.1)
                    q1.queue.clear()
                    
                    
                if len(data)>=self.dump_threshold:
                    self.write2xl(data)
                    
                    data=[]
                    
                    
            else:
                # print("waiting to print")
                pass

            if event.is_set():
                break
            sleep(0.5)

    def write2xl(self,dataList):
        #1 . Open xl
        count = self.openxl()
        row_count =count+1
        buy=0
        sell=0
        #2. Save
        for _row, data in enumerate(dataList):
            for _col,_d in enumerate(data):
                self.ws.cell(row=row_count+_row, column=1+_col).value = _d
            
            # success=data["success"]
            # depth = success["depth"]
            # buy= [(float(depth[0]["buy"][i]['price'])* float(depth[0]["buy"][i]['quantity']))for i in range(0,5)]
            # sell= [(float(depth[0]["sell"][i]['price'])* float(depth[0]["sell"][i]['quantity']))for i in range(0,5)]
            # buy = sum(buy)
            # sell = sum(sell)
            
            # if ((buy+sell)== 0): # condition 1
            #     continue
            
            # for _col,_d in enumerate([buy, sell,depth[0]["upper_ckt_limit"], depth[0]["lower_ckt_limit"],datetime.now()]):
            #     self.ws.cell(row=row_count+_row, column=1+_col).value = _d
            

        #3. Save the changes
        self.wb.save(self.csv_dump_path)
        
        #4. close
        self.closexl()
        
        
        
        
        
    def closexl(self,):
        try:
            self.wb.close()
        except Exception as e:
            print(f"closing excel: {e}")
            
            
    def createXLXS(self,):
        if not (os.path.exists(self.csv_dump_path)):
            col_heading=["id","token","Best buy price","Best buy quantity","Best sell price","Best sell quantity",
            "Last trade price","High price","Low price","Average trade price","Closing price","Open price",
            "Net change percentage","Total sell quantity","Total buy quantity","Total trade quantity","Open Interest",
            "Total trade value","Last trade quantity","Last trade time","Net change","Upper circuit limit","Lower circuit limit"]
            df =pd.DataFrame(columns=col_heading)
            df
            wb =  Workbook()
            ws = wb['Sheet']

            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            wb.save(self.csv_dump_path)
            print()
            
            
    def openxl(self,):            
        try:
            self.wb = load_workbook(filename=self.csv_dump_path)
            self.ws = self.wb['Sheet']
        except Exception as e:
            self.wb = Workbook()
            self.ws = self.wb.active
            
            
        return self.ws.max_row
            
        
        
            
        
            
        
class SAVE_DATA():
    def __init__(self) -> None:
        pass

    def print_q(self,):
        while (True):
            if (self.q_print > 0):
                if(self.q_print == 2):
                    for i in q2:
                        print(i)
                else:
                    for i in q1:
                        print(i)
            else:
                print("waiting to print")

            if event.is_set():
                break
            sleep(0.5)


if __name__ == "__main__":
    e = DUMMY_KOTAK()
    e.kk.kotak_login()
    # e.kk.get_history()
    # e.kk.get_report()

    e.kk.kotat_subscribe_(e.stock_call_back)
    
    # e.kk.kotat_subscribe_(e.send_fake_stock)

    # t = Thread(target=e.callBackStock, args=())
    t2 = Thread(target=e.print_q, args=())

    # t.start()
    t2.start()

    time430pm= datetime.time( 16,30,0 ) # Time, without a date
    
    while True:
        try:
            if ( time430pm < datetime.datetime.now().time()):
                break
            else:
                sleep(1)
        except KeyboardInterrupt:
            event.set()
            break
        
        
    event.set()

    # t.join()
    t2.join()
    if not q1.empty():
        print(q1.get())
    if not q2.empty():
        print(q2.get())
